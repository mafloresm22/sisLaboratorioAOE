import os
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpImage
from utils.paths import get_resource_path

def exportar_instrumentos_excel(items: list, filepath: str, responsable_nombre: str = "Usuario Actual"):
    wb = openpyxl.Workbook()
    
    # Ruta del logo
    logo_path = get_resource_path(os.path.join("assets", "img", "AOE_img.jpeg"))
    
    # Remover la hoja por defecto que crea openpyxl
    if "Sheet" in wb.sheetnames:
        default_sheet = wb["Sheet"]
        wb.remove(default_sheet)
        
    # Agrupar los items por 'Ubicación' 
    grouped_items = defaultdict(list)
    for item in items:
        # Obtenemos la ubicación, asumiendo el campo 'nombre_laboratorio'
        ubic = getattr(item, "nombre_laboratorio", "Sin Ubicación")
        if not ubic:
            ubic = "Sin Ubicación"
        grouped_items[ubic].append(item)
        
    # --- ESTILOS DE CELDAS ---
    title_fill  = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    alt_fill    = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
    
    title_font    = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    subtitle_font = Font(name="Arial", size=11, italic=True, color="FFFFFF")
    header_font   = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font     = Font(name="Arial", size=10)
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color="BDC3C7"),
        right=Side(style='thin', color="BDC3C7"),
        top=Side(style='thin', color="BDC3C7"),
        bottom=Side(style='thin', color="BDC3C7")
    )

    # Definición y anchos de columnas
    columns = [
        ("N°", 8),
        ("DESCRIPCIÓN", 35),
        ("CANT.", 12),
        ("MARCA", 15),
        ("MODELO", 15),
        ("SERIE", 15),
        ("COLOR", 12),
        ("TAMAÑO", 12),
        ("ESTADO CONS.", 15),
        ("UBICACIÓN", 20),
        ("RESPONSABLE", 20),
        ("PISO", 10)
    ]
    
    sheet_counter = 1
    
    # Lógica de títulos dinámicos e info de fecha
    current_year = datetime.now().year
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    
    for ubic, sub_items in grouped_items.items():
        # Procesar en bloques de 30 registros máximos por hoja
        for i in range(0, len(sub_items), 30):
            chunk = sub_items[i:i+30]
            
            # Crear nombre para la hoja de Excel 
            safe_ubic = str(ubic).translate(str.maketrans("", "", r"[]:*?/\\"))
            sheet_name = f"{safe_ubic[:20]}_{sheet_counter}"
            
            ws = wb.create_sheet(title=sheet_name)
            sheet_counter += 1
            
            # Configurar el ancho de las columnas
            for col_idx, (col_name, width) in enumerate(columns, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            
            # === LOGO Y TÍTULOS ===
            ws.row_dimensions[1].height = 45
            ws.row_dimensions[2].height = 30
            
            if os.path.exists(logo_path):
                try:
                    img = OpImage(logo_path)
                    img.width = 55
                    img.height = 55
                    ws.add_image(img, 'A1')
                except Exception as e:
                    print(f"⚠️ No se pudo cargar el logo en Excel: {e}")

            # Lógica de títulos dinámicos
            ubic_norm = str(ubic).upper()
            es_especial = any(k in ubic_norm for k in ["BIOLOGIA", "BIOLOGÍA", "QUIMICA", "QUÍMICA", "FISICA", "FÍSICA"])
            es_ecologia = "ECOLOGIA" in ubic_norm or "ECOLOGÍA" in ubic_norm

            if es_especial:
                titulo_principal = f'INVENTARIO DE LA INSTITUCIÓN EDUCATIVA "ANTENOR ORREGO ESPINOZA" - {current_year}'
                subtitulo = f"SECTOR ENFOQUE: {ubic_norm} | FECHA: {fecha_actual}"
            elif es_ecologia:
                titulo_principal = 'INVENTARIO FÍSICO DE BIENES MUEBLES DEL "AULA ECOLOGICA"'
                subtitulo = f"REPORTE DETALLADO DE ACTIVOS | FECHA: {fecha_actual}"
            else:
                titulo_principal = f"INVENTARIO DE INSTRUMENTOS - {ubic_norm}"
                subtitulo = f"REPORTE DE ESTADO | FECHA: {fecha_actual}"

            # === RENDERIZAR TÍTULOS ===
            ws.merge_cells('A1:L1')
            ws['A1'] = titulo_principal
            ws['A1'].font = title_font
            ws['A1'].fill = title_fill
            ws['A1'].alignment = center_align
            
            ws.merge_cells('A2:L2')
            ws['A2'] = subtitulo
            ws['A2'].font = subtitle_font
            ws['A2'].fill = title_fill
            ws['A2'].alignment = center_align
            
            # === CABECERAS DE TABLA ===
            header_row = 4
            for col_idx, (col_name, _) in enumerate(columns, 1):
                cell = ws.cell(row=header_row, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
                
            # === DATOS DE LA TABLA ===
            current_row = header_row + 1
            for item in chunk:
                # Obtención segura de atributos del instrumento
                numero      = str(getattr(item, "idInstrumento", "-")).zfill(2)
                descripcion = getattr(item, "descripcionInstrumento", "") or "Sin descripción"
                cantidad    = getattr(item, "cantidadInstrumento", 0)
                unidad      = getattr(item, "nombre_unidad", "und.")
                cant_str    = f"{cantidad:g} {unidad}"
                marca       = getattr(item, "marcaInstrumento", "") or "—"
                modelo      = getattr(item, "modeloInstrumento", "") or "—"
                serie       = getattr(item, "serieInstrumento", "") or "—"
                color       = getattr(item, "colorInstrumento", "") or "—"
                tamano      = getattr(item, "tamanoInstrumento", "") or "—"
                estado      = str(getattr(item, "idEstadoCons", "—"))
                piso        = getattr(item, "pisoInstrumento", "") or "—"
                
                row_data = [
                    numero, descripcion, cant_str, marca, modelo, 
                    serie, color, tamano, estado, ubic, 
                    responsable_nombre, piso
                ]
                
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.font = data_font
                    cell.border = thin_border
                    
                    # N°, Cantidad, Estado y Piso centrados. Lo demás, alineado a la izquierda
                    if col_idx in [1, 3, 9, 12]:  
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align
                        
                    # Aplicar color a filas alternas para "formato tabla"
                    if (current_row - header_row) % 2 != 0:
                        cell.fill = alt_fill
                        
                current_row += 1
                
            # === APARTADO DE FIRMA ===
            firma_row = current_row + 6  
            
            ws.merge_cells(start_row=firma_row, start_column=5, end_row=firma_row, end_column=8)
            ws.merge_cells(start_row=firma_row+1, start_column=5, end_row=firma_row+1, end_column=8)
            ws.merge_cells(start_row=firma_row+2, start_column=5, end_row=firma_row+2, end_column=8)

            ws.cell(row=firma_row, column=5).value = "________________________________________"
            ws.cell(row=firma_row, column=5).alignment = center_align
            
            ws.cell(row=firma_row+1, column=5).value = f"Responsable: {responsable_nombre}"
            ws.cell(row=firma_row+1, column=5).font = Font(name="Arial", size=11, bold=True)
            ws.cell(row=firma_row+1, column=5).alignment = center_align
            
            ws.cell(row=firma_row+2, column=5).value = "Firma del Docente / Responsable"
            ws.cell(row=firma_row+2, column=5).font = data_font
            ws.cell(row=firma_row+2, column=5).alignment = center_align

    # Si por alguna razón la lista venía vacía
    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet("Sin Datos")
        ws['A1'] = "No se encontraron registros para exportar."

    wb.save(filepath)
